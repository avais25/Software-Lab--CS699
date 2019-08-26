import requests
import sys
import re
import csv
#import pandas as pd
import openpyxl


def main(argv):
    workNo=sys.argv[2]			#initial variables
    sheetIndex=0
    foundFlag=0
    headerFlag=0
    fileName=sys.argv[1]
    fd = open(fileName)
    fdReader=csv.reader(fd)
    csvData = list(fdReader)
    print("Initial Data in csv \n",csvData)
    
    wb = openpyxl.Workbook()

    
    for row in csvData:
      for ele in row:
        if re.search('work no', ele, re.IGNORECASE):		#getting index
         rowIndex=csvData.index(row)
         colIndex=row.index(ele)
         headerFlag=1
         #print(rowIndex)
         
         print("\nSubstring \'work no\' matched with header:-  ",ele)
         print("\n Index of matched header:-  ",colIndex)


    if(headerFlag==1):			#checking if header exist
     for row in csvData:
       if row[colIndex]==workNo:
        foundFlag=1
        newTitle='%s_%d'% (workNo,(sheetIndex+1))
        wb.create_sheet(index=sheetIndex, title=newTitle)		#creating sheets
        sheet = wb.get_sheet_by_name(newTitle)
        #sheet['A1'] = 'Hello world!'
        sheet.append(csvData[0])
        newRow=row
        newRow[colIndex]='%s_%d'% (workNo,(sheetIndex+1))
        sheet.append(row)
        
        #wb.create_sheet(index=0, title='First Sheet')
        #wb.create_sheet(index=0, title='Second Sheet')
        sheetIndex=sheetIndex+1
  
        print("\nMatch Number ",sheetIndex)
        print(row)
        
     if(foundFlag==1):				#checking if work no exist		
      for sheet in wb.worksheets:
        #print(sheet)
        if(sheet.title=="Sheet"):
         wb.remove_sheet(sheet)
      outFileName=re.sub(r'.csv','_%s.xlsx'%workNo,fileName)
      print("\nOutput file is:-",outFileName)
      wb.save(outFileName)			#saving output file
     else:
      print("\n No record Matched the work no.\nOutput file will not be created.")
    else:
      print("\n No header conatin the substring  \'work no\'")
    return




if __name__ == "__main__":
    main(sys.argv)
